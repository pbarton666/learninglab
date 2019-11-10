import os
import pandas as pd
from openpyxl import load_workbook

data_dir = "."
template = 'chartme_template.xlsx'
new_wkbk = 'chartme_data_added.xlsx'
tab_name = 'data'

def create_data():
    """Creates and returns a DataFrame"""
    data = \
        [
          ['baseball', 180, 1100],
          ['wrestling', 30,  300],
          ['gymnastics', 1,  120],      
        ]
    cols = ['sport', 'duration', 'fans' ]
    
    sports_df = pd.DataFrame(data=data, columns=cols)
    
    return sports_df
def insert_data_into_template(df, template='', tab_name='', new_book_name=''):
    """Inserts the contents of a DataFrame object into a tab of an existing
           spreadsheet and saves the result."""
    
    #Uses pandas' default Excel engine openpyxl directly.
    template_book = load_workbook(template)
    template_data_wksh = template_book[tab_name]
    
    #We can write the template a cell at a time.
    #... the first line will be the columns of the DataFrame
    for col, col_heading in enumerate(df.columns, 2):
        #Returns a Cell object, but we don't need it
        template_data_wksh.cell(row=1, column=col, value = col_heading)
        
    #Write the contents of the DataFrame (note the spreadsheet is "1-based")    
    rows, cols = df.shape
    for out_row, df_row in enumerate(range(rows), 2):
        this_row = df.iloc[df_row, :]
        for out_col, df_col in enumerate(range(cols), 1):      
            cell_obj = template_data_wksh.cell(row=out_row, column=out_col, 
                                               value = this_row[df_col])
            
    #This closes the openpyxl workbook object and saves it as a file.
    template_book.save(new_book_name)       
                      
df = create_data()  
template = os.path.abspath(os.path.join(data_dir,  template))
new_book_name = os.path.abspath(os.path.join(data_dir, new_wkbk))

insert_data_into_template(df,  template=template, tab_name=tab_name,
                               new_book_name=new_book_name)

print("Updated template is created: \n\t{}".format(new_book_name))
a=1





